import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

workbook = openpyxl.load_workbook('example.xlsx')
print(workbook.sheetnames) # The workbook's sheets' names.
sheet = workbook['Sheet1']
print(sheet['A1']) # Get a cell from the sheet.
print(sheet['A1'].value) # Get the value from the cell.

print(sheet['B1'].value) # Get another cell from the sheet.
c = sheet['B1']
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))

print(sheet['C1'].value)

# getting cells by numeric column and row values

print(sheet.cell(row=1, column=2)) # returns cells B1
print(sheet.cell(row=1, column=2).value)

for i in range (1, 8, 2):
    print(i, sheet.cell(row=i, column= 2).value) #print every other row in column 2

# get max rows and max columns of the worksheet
print(sheet.max_row)
print(sheet.max_column) # is an integrer and not a letter (A-Z)
