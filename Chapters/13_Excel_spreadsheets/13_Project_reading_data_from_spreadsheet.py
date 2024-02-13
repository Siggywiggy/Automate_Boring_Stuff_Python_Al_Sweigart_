#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for
# each county.

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pprint

# Open and read the cells of an Excel document with the openpyxl module.

workbook = openpyxl.load_workbook('censuspopdata.xlsx')
print(workbook.sheetnames)

sheet = workbook[workbook.sheetnames[0]]

print(sheet)
print(sheet.max_column)
print(sheet.max_row)
print(get_column_letter(sheet.max_column))

county_data = {}

# TODO: Fill in countyData with each county's population and tracts.

print('Reading rows...')

# Calculate all the tract and population data and store it in a data structure.

for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    # Make sure the key for this state exists.
    county_data.setdefault(state, {})
    # Make sure the key for this county in this state exists.
    county_data[state].setdefault(county, {'tracts' : 0, 'pop' : 0})
    # Each row represents one census tract, so increment by one.
    county_data[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    county_data[state][county]['pop'] += int(pop)


# TODO: Open a new text file and write the contents of countyData to it.

print('Writing results...')
result_file = open('census2010.py', 'w')
result_file.write('all_data = ' + pprint.pformat(county_data))
result_file.close()
print('Done!')