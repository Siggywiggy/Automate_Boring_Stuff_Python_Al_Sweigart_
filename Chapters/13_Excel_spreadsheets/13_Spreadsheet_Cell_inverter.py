#! python3
# a program to invert the rows and cells in a spreadsheet
import openpyxl
import logging
from openpyxl.utils import get_column_letter, column_index_from_string

# comment out to enable logging
logging.disable(logging.CRITICAL)

logging.basicConfig(
    level=logging.DEBUG, format=" %(asctime)s -  %(levelname) s -  %(message)s"
)

logging.debug("Start of program")

# load the input data file
workbook_data = openpyxl.load_workbook('sample_chart.xlsx')
sheet_data = workbook_data.active
#create workbook and sheet objects for output file
workbook_output = openpyxl.Workbook()
sheet_output = workbook_output.active

sheet_data_list = list()

# iterate over all the rows of cell objects in input data sheet
for row_of_cell_objects in sheet_data['A1' : get_column_letter(sheet_data.max_column) + str(sheet_data.max_row)]:
    logging.debug(f'current row is {tuple(row_of_cell_objects)}')
    # for every cell
    for cell in row_of_cell_objects:
        logging.debug(f'column letter is {cell.column_letter}, row is {cell.row}')
        # append a list with content of the cell, the column index and the cell row index
        sheet_data_list.append([cell.value, column_index_from_string(cell.column_letter), cell.row])

for sub_list in sheet_data_list:
    # write the data to output data, inverting the column and row indexes
    sheet_output[get_column_letter(sub_list[2]) + str(sub_list[1])] = sub_list[0]

# save the new workbook to a xlsx file
workbook_output.save('Inverted_cells.xlsx')