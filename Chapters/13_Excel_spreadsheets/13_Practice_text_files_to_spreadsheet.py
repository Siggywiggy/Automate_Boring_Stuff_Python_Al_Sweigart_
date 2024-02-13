#! python3 a program that reads the contents of several text files
# the lines of the first text file will be in column A
# lines in the second text file will be in column B

import openpyxl
import logging
from openpyxl.utils import get_column_letter

# comment out to enable logging
#logging.disable(logging.CRITICAL)

logging.basicConfig(
    level=logging.DEBUG, format=" %(asctime)s -  %(levelname) s -  %(message)s"
)

logging.debug("Start of program")

#create workbook and sheet objects for output file
workbook_output = openpyxl.Workbook()
sheet = workbook_output.active

input_files = ['input_file_1.txt', 'input_file_2.txt', 'input_file_3.txt']

for i in range(0, len(input_files)):
    # each file in a column, using their order in list to determine column letter
    column_letter = get_column_letter(i+1)
    logging.debug(f'column letter is {column_letter}')
    # open the nth input file in the list
    input_file = open(input_files[i], 'r')
    # create a list of rows from the .readlines() method
    input_list = input_file.readlines()
    logging.debug(f'input list is {input_list}')
    # loop over all the rows with enumerate to also get the row index
    for index, line in enumerate(input_list):
        logging.debug(f"Line {index}: {line}")
        sheet[column_letter + str(index + 1)] = line


# save the new workbook to a xlsx file
workbook_output.save('Text_files_to_columns.xlsx')