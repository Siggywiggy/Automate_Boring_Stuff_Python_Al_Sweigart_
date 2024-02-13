import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook('produceSales.xlsx')

print(workbook.sheetnames)

sheet = workbook.worksheets[0]

price_updates = {
    'Garlic' : 3.07,
    'Celery' : 1.19,
    'Lemon' : 1.27
}


print(f'minimum column is {sheet.min_column}')

for row_of_cell_objects in sheet[get_column_letter(sheet.min_column) + str(sheet.min_row) : get_column_letter(sheet.max_column) + str(sheet.max_row)]:
    if row_of_cell_objects[0].value in price_updates.keys():
        sheet[row_of_cell_objects[1].coordinate].value = price_updates[row_of_cell_objects[0].value]


workbook.save('updated_produce_sales.xlsx')