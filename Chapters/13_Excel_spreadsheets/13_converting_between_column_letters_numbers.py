import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

workbook = openpyxl.load_workbook('example.xlsx')
print(workbook.sheetnames)

sheet = workbook['Sheet1']
print(get_column_letter(sheet.max_column)) # whats the max existing columns letter

print(column_index_from_string('A')) #get A's numberic value
print(column_index_from_string('AA'))