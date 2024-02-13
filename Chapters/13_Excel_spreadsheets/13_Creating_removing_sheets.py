import openpyxl

workbook = openpyxl.Workbook()

print(workbook.sheetnames)

workbook.create_sheet()

print(workbook.sheetnames)

# Create a new sheet at index 0.

workbook.create_sheet(index=0, title='First Sheet')
print(workbook.sheetnames)

workbook.create_sheet(index=2, title='Middle Sheet')
print(workbook.sheetnames)

del workbook['Sheet1']
print(workbook.sheetnames)