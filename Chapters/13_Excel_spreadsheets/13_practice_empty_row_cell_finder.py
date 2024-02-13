import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


workbook = openpyxl.load_workbook('empty_row.xlsx')

sheet = workbook.active
empty_cells = list()
empty_rows = list()

max_columns = sheet.max_column
print(f'Max columns are {max_columns}')
max_column_letter_index = get_column_letter(sheet.max_column)
print(f'Max column letter is {max_column_letter_index}')
max_row = sheet.max_row
print(f'Max rows are {max_row}')

# create a slice of the whole data in workbook and use it in a for loop

for row_of_cell_objects in sheet['A1' : max_column_letter_index + str(max_row)]:
    empty_cells_in_row = 0
    print(list(row_of_cell_objects))
    for cell_object in row_of_cell_objects:
        if cell_object.value is None:
            empty_cells_in_row += 1
            empty_cells.append(cell_object.coordinate)
    if empty_cells_in_row >= 3:
        empty_rows.append(row_of_cell_objects)

print('Empty cells are: ')
print(empty_cells)
print('Empty rows are: ')
print(list(empty_rows))