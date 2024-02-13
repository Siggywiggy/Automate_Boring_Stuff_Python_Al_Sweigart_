import openpyxl
import logging
from openpyxl.utils import get_column_letter

logging.disable(logging.CRITICAL)

logging.basicConfig(
    level=logging.DEBUG, format=" %(asctime)s -  %(levelname) s -  %(message)s"
)

logging.debug("Start of program")


# blank row starting row and how many blank rows
blank_rows_start = 50
blank_rows_count = 10
# load the input data file
workbook_data = openpyxl.load_workbook('censuspopdata.xlsx')
sheet_data = workbook_data.active
#create workbook and sheet objects for output file
workbook_output = openpyxl.Workbook()
sheet_output = workbook_output.active

#iterate over all rows in input workbook sheet, enumerate to keep count of the rows

for i, row_of_cells in enumerate(sheet_data.iter_rows(max_row=sheet_data.max_row, max_col=sheet_data.max_column)):
    if i < blank_rows_start: # copy the value over to the exact same cell
        for cell in row_of_cells:
            logging.debug(f'the current cell value is {cell.value}, i is < {blank_rows_start} (blank rows start row)')
            logging.debug(f'cell coordinate is {cell.coordinate}')
            sheet_output[cell.coordinate] = cell.value

    else: # copy the value over to a shifted by blank_rows_count shell
        for cell in row_of_cells:
            logging.debug(f'the current cell value is {cell.value}, i is > {blank_rows_start} (blank rows start row)')
            new_cell_row = cell.row + blank_rows_count
            cell_column = cell.column
            logging.debug(f'new shifted row is {new_cell_row}, current column is {cell_column}')
            sheet_output[get_column_letter(cell_column) + str(new_cell_row)] = cell.value

# save the new workbook to a xlsx file
workbook_output.save('Blank_rows.xlsx')