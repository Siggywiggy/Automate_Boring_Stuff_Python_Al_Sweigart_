import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
# Asheet.unmerge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5') # Merge these two cell
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')