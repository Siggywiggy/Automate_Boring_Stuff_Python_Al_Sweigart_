import openpyxl

workbook = openpyxl.Workbook() # create a blank workbook
print(workbook.sheetnames) # starts with a singular worksheet
sheet = workbook.active # create a sheet object
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet' # changing the title
print(workbook.sheetnames)

workbook.create_sheet()
print(workbook.sheetnames)


workbook.save('example_copy.xlsx')