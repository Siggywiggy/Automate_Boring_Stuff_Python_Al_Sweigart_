#! python 3
# a program takes a number (N) from command line and creates a NxN multiplication table in excel workbook

import sys
import openpyxl
from openpyxl.styles import Font

#take the number argument from command line
multiplication_number = int(sys.argv[1])

# open workbook, create a sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# create a font style for labels
label_font = Font(name='Helvetica', bold=True, size=12)

# fill table with labels
# row labels in cells B1 - > N1 - range(1, n+1)
for i in range(1, multiplication_number + 1):
    # row labels in cells B1 - > N1 range(1, n+1)
    column_letter = openpyxl.utils.cell.get_column_letter(i+1)
    sheet[column_letter + '1'] = i
    sheet[column_letter + '1'].font = label_font
    # column labels in column A down to row N + 1
    sheet['A' + str(i+1)] = i
    sheet['A' + str(i + 1)].font = label_font

# start multiplying from B2 to max_row, max_column
for i in range(2, sheet.max_column + 1): # shift the range to compensate the gap at A1
    for y in range(2, sheet.max_row + 1):
        sheet[openpyxl.utils.cell.get_column_letter(i)+str(y)] = (i-1) * (y-1)

# save the excel workbook

workbook.save('Multiplication_table.xlsx')

