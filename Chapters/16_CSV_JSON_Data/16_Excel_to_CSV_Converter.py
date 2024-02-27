import csv
import sys
import openpyxl
import os
import logging
from pathlib import Path

# comment out to enable logging
#logging.disable(logging.CRITICAL)

logging.basicConfig(
    level=logging.DEBUG, format=" %(asctime)s -  %(levelname) s -  %(message)s"
)

logging.debug("Start of program")

starting_folder = Path.cwd() / 'excelSpreadsheets'
logging.debug(f'the starting location is {starting_folder}')

# loop trough every file matching the glob pattern in folder
for file_name in list(starting_folder.glob("*.xlsx")):
    logging.debug(f'file name is {file_name}')
    workbook = openpyxl.load_workbook(file_name)
    # loop trough every sheet in the workbook
    for sheet in workbook.sheetnames:
        logging.debug(f'sheet name is {sheet}')
        # create the CSV filename from Excel filename and sheet title
        csv_filename = starting_folder / (os.path.basename(file_name).split('.')[0] + '_' + sheet + '.' + os.path.basename(file_name).split('.')[1])
        logging.debug(f'new csv file name is {csv_filename}')
        # Create the csv.writer object for this CSV file.
        output_file = open(csv_filename, 'w', newline='')
        output_writer_object = csv.writer(output_file)
        # loop trough every row in the sheet
        for row_num in range(1, sheet.max_row +1):
            row_data = list()
            for col_num in range(sheet.max_column +1):
                row_data.append(sheet.cell(row=row_num, column=col_num).value)

            print(row_data)

